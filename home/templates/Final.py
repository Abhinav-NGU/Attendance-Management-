import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import openpyxl


# Function to preprocess images before encoding
def preprocess_image(img):
    img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)  # Convert to grayscale
    img = cv2.equalizeHist(img)  # Apply histogram equalization
    return img

def findEncodings(images):
    encodeList = []
    for img in images:
        img = preprocess_image(img)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        face_encodings = face_recognition.face_encodings(img)
        if len(face_encodings) > 0:
            encodeList.append(face_encodings[0])
    return encodeList

def run_face_recognition(image_data):
    path = 'home/templates/ImagesAttendance'
    images = []
    classNames = []
    encodeListKnown = []

    myList = os.listdir(path)

    for c1 in myList:
        curImg = cv2.imread(f'{path}/{c1}')
        images.append(curImg)
        classNames.append(os.path.splitext(c1)[0])

    print(classNames)
    encodeListKnown = findEncodings(images)
    print('Encoding Complete')
    i=1
    # Load the image on which you want to perform face recognition
    image_to_recognize = cv2.imread(image_data)
    i=i+1  # Replace 'path_to_image.jpg' with the path to your image
    image_to_recognize = cv2.cvtColor(image_to_recognize, cv2.COLOR_BGR2RGB)

    # Recognize faces in the loaded image
    facesCurFrame = face_recognition.face_locations(image_to_recognize)
    encodesCurFrame = face_recognition.face_encodings(image_to_recognize, facesCurFrame)

    # Create an Excel file with a name in the format "Attendance - YYYY-MM-DD.xlsx"
    now = datetime.now()
    excel_file_name = f'templates/Attendance.xlsx'
    ws_title = "Attendance"

    # Check if the Excel file already exists
    if os.path.exists(excel_file_name):
        wb = openpyxl.load_workbook(excel_file_name)
        # Check if the sheet already exists
        if ws_title in wb.sheetnames:
            ws = wb[ws_title]
        else:
            ws = wb.create_sheet(title=ws_title)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ws_title
        # Write the headings "RegNo," "Name," "Date," "Time," and "Accuracy" if creating a new sheet
        ws.cell(row=1, column=1, value="RegNo")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Date")
        ws.cell(row=1, column=4, value="Time")
        ws.cell(row=1, column=5, value="Accuracy")

    # Inside the loop where faces are recognized
    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace, tolerance=0.6)  # Experiment with the tolerance value
        faceDistances = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDistances)

        accuracy = (1 - faceDistances[matchIndex]) * 100  # Convert distance to accuracy percentage

        if matches[matchIndex] and accuracy >= 50:  # Set the minimum accuracy threshold here (e.g., 50)
            name_with_reg = classNames[matchIndex].upper()
            name_reg_parts = name_with_reg.split('-')
            
            # Separate name and registration number
            if len(name_reg_parts) == 2:
                registration_number = name_reg_parts[1]
                student_name = name_reg_parts[0]

                y1, x2, y2, x1 = faceLoc
                cv2.rectangle(image_to_recognize, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(image_to_recognize, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(image_to_recognize, f'{student_name}', (x1 + 6, y2 - 6), 
                            cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)

                # Check if the registration number already exists in the sheet
                reg_no_col = 1  # Assuming column 1 contains registration numbers
                reg_no_row = 2  # Starting from row 2 (assuming row 1 contains headings)
                while ws.cell(row=reg_no_row, column=reg_no_col).value is not None:
                    if ws.cell(row=reg_no_row, column=reg_no_col).value == registration_number:
                        # Update the data for the existing registration number
                        ws.cell(row=reg_no_row, column=1, value=student_name)
                        ws.cell(row=reg_no_row, column=3, value=now.strftime('%Y-%m-%d'))
                        ws.cell(row=reg_no_row, column=4, value=now.strftime('%H:%M:%S'))
                        ws.cell(row=reg_no_row, column=5, value=accuracy)
                        break
                    reg_no_row += 1
                else:
                    # If the registration number is not found, append the data to the next empty row
                    empty_row = reg_no_row
                    ws.cell(row=empty_row, column=2, value=registration_number)
                    ws.cell(row=empty_row, column=1, value=student_name)
                    ws.cell(row=empty_row, column=3, value=now.strftime('%Y-%m-%d'))
                    ws.cell(row=empty_row, column=4, value=now.strftime('%H:%M:%S'))
                    ws.cell(row=empty_row, column=5, value=accuracy)

                # Capture the timestamp
                now = datetime.now()
                dtString = now.strftime('%Y-%m-%d %H:%M:%S')
                print(f'{student_name} recognized at {dtString}')

    # Save the Excel file
    wb.save(excel_file_name)
    resized_image = cv2.resize(image_to_recognize, (300, 300))

# Display the resized image
#cv2.imshow('Resized Image with Recognized Faces', resized_image)
# Display the image with recognized faces
# cv2.imshow('Image with Recognized Faces', image_to_recognize)
# cv2.waitKey(0)
# cv2.destroyAllWindows()
